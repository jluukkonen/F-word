import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sklearn.metrics import cohen_kappa_score, confusion_matrix, classification_report
import os

# ── Paths ──
HUMAN_FILE = "/Volumes/United/Work/F-word/swearing-nlp/data/labeled/to_label.xlsx"
CHATGPT_FILE = "/Volumes/United/Work/F-word/swearing-nlp/data/labeled/chatgpt_labels.tsv"
GEMINI_FILE = "/Volumes/United/Work/F-word/swearing-nlp/data/labeled/gemini_labels.tsv"
FRIEND_FILE = "/Volumes/United/Work/F-word/swearing-nlp/data/labeled/friend1_labels.tsv"
FRIEND2_FILE = "/Volumes/United/Work/F-word/swearing-nlp/data/labeled/friend2_labels.tsv"
FRIEND3_FILE = "/Volumes/United/Work/F-word/swearing-nlp/data/labeled/friend3_labels.tsv"
OUTPUT_FILE = "/Volumes/United/Work/F-word/swearing-nlp/results/interrater_analysis.xlsx"

# ── Styles ──
DARK_BLUE = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
MID_BLUE = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
DARK_GREEN = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
DARK_RED = PatternFill(start_color='B71C1C', end_color='B71C1C', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
LIGHT_GREEN = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
LIGHT_RED = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
ZEBRA = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

HEADER_FONT = Font(color='FFFFFF', bold=True, size=11, name='Arial')
TITLE_FONT = Font(color='1F4E78', bold=True, size=16, name='Arial')
SUBTITLE_FONT = Font(color='1F4E78', bold=True, size=12, name='Arial')
BODY = Font(size=10, name='Arial')
BOLD = Font(size=10, name='Arial', bold=True)
BIG_NUM = Font(size=22, name='Arial', bold=True, color='1F4E78')
SMALL_LABEL = Font(size=9, name='Arial', color='666666', bold=True)

BSIDE = Side(style='thin', color='BFBFBF')
BORDER = Border(left=BSIDE, right=BSIDE, top=BSIDE, bottom=BSIDE)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

LABEL_FILLS = {
    'aggression': PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid'),
    'bonding': PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'),
    'emphasis': PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid'),
    'frustration': PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid'),
    'ambiguous': PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid'),
}

def hdr(ws, row, ncols, fill=DARK_BLUE):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = HEADER_FONT; cell.alignment = CENTER; cell.border = BORDER

def dcell(cell, i, align=CENTER):
    cell.font = BODY; cell.alignment = align; cell.border = BORDER
    if i % 2 == 0: cell.fill = ZEBRA

def lcol(cell, label):
    lbl = str(label).strip().lower()
    if lbl in LABEL_FILLS: cell.fill = LABEL_FILLS[lbl]

def write_title(ws, text, row=1, merge_to='H'):
    ws.merge_cells(f'A{row}:{merge_to}{row}')
    c = ws[f'A{row}']; c.value = text; c.font = TITLE_FONT
    ws.row_dimensions[row].height = 40

def write_subtitle(ws, text, row, col=1):
    c = ws.cell(row=row, column=col); c.value = text; c.font = SUBTITLE_FONT
    ws.row_dimensions[row].height = 28

def write_cm(ws, start_row, cm_data, labels_list, title, annotator_a, annotator_b):
    """Write a confusion matrix block starting at start_row. Returns next available row."""
    write_subtitle(ws, title, start_row)
    r = start_row + 1
    # Corner
    corner = ws.cell(row=r, column=1)
    corner.value = f"{annotator_a} ↓  {annotator_b} →"
    corner.font = Font(color='FFFFFF', bold=True, size=9, name='Arial')
    corner.fill = PatternFill(start_color='37474F', end_color='37474F', fill_type='solid')
    corner.border = BORDER
    # Col headers
    for j, lbl in enumerate(labels_list):
        c = ws.cell(row=r, column=j+2); c.value = lbl
        c.fill = MID_BLUE; c.font = HEADER_FONT; c.alignment = CENTER; c.border = BORDER
    # Rows
    for i, lbl in enumerate(labels_list):
        rr = r + 1 + i
        rh = ws.cell(row=rr, column=1); rh.value = lbl; rh.font = BOLD; rh.border = BORDER
        lcol(rh, lbl)
        for j in range(len(labels_list)):
            c = ws.cell(row=rr, column=j+2); val = int(cm_data[i][j]); c.value = val
            c.alignment = CENTER; c.border = BORDER; c.font = BODY
            if i == j:
                c.fill = PatternFill(start_color='A5D6A7', end_color='A5D6A7', fill_type='solid')
                c.font = Font(size=12, name='Arial', bold=True, color='1B5E20')
            elif val >= 5:
                c.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                c.font = Font(size=10, name='Arial', bold=True, color='B71C1C')
            elif val >= 2:
                c.fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
    return r + 1 + len(labels_list) + 1

def main():
    # ── Load ──
    human = pd.read_excel(HUMAN_FILE).rename(columns={"label": "human"})
    chatgpt = pd.read_csv(CHATGPT_FILE, sep="\t").rename(columns={"label": "chatgpt"})
    gemini = pd.read_csv(GEMINI_FILE, sep="\t").drop_duplicates(subset="id", keep="first").rename(columns={"label": "gemini"})
    friend = pd.read_csv(FRIEND_FILE, sep="\t").rename(columns={"label": "friend1"})
    friend2 = pd.read_csv(FRIEND2_FILE, sep="\t").rename(columns={"label": "friend2"})
    friend3 = pd.read_csv(FRIEND3_FILE, sep="\t").rename(columns={"label": "friend3"})

    # Clean
    for df, col in [(human, "human"), (chatgpt, "chatgpt"), (gemini, "gemini"), (friend, "friend1"), (friend2, "friend2"), (friend3, "friend3")]:
        df[col] = df[col].astype(str).str.strip().str.lower()
    chatgpt["chatgpt"] = chatgpt["chatgpt"].replace("neutral", "ambiguous")

    # Merge all three
    merged_all = human.merge(gemini, on="id", how="inner")
    merged_all = merged_all[merged_all["human"].notna() & (merged_all["human"] != "") & (merged_all["human"] != "nan")]
    merged_all = merged_all[merged_all["gemini"].notna() & (merged_all["gemini"] != "") & (merged_all["gemini"] != "nan")]

    # 3-way: also has ChatGPT
    three_way = merged_all.merge(chatgpt, on="id", how="inner")
    three_way = three_way[three_way["chatgpt"].notna() & (three_way["chatgpt"] != "") & (three_way["chatgpt"] != "nan")]

    labels = sorted({"aggression", "bonding", "emphasis", "frustration", "ambiguous"})

    # ── Compute stats ──
    def pair_stats(df, cola, colb):
        n = len(df)
        agree = (df[cola] == df[colb]).sum()
        kappa = cohen_kappa_score(df[cola], df[colb], labels=labels)
        cm = confusion_matrix(df[cola], df[colb], labels=labels)
        return n, agree, kappa, cm

    # Human vs Gemini (full overlap)
    hg_n, hg_agree, hg_kappa, hg_cm = pair_stats(merged_all, "human", "gemini")
    # Human vs ChatGPT (93 overlap)
    hc_n, hc_agree, hc_kappa, hc_cm = pair_stats(three_way, "human", "chatgpt")
    # Gemini vs ChatGPT (93 overlap)
    gc_n, gc_agree, gc_kappa, gc_cm = pair_stats(three_way, "gemini", "chatgpt")
    
    # Human vs Friend (50 overlap)
    # Filter human for 50 items
    merged_friend = human.merge(friend, on="id", how="inner")
    hf_n, hf_agree, hf_kappa, hf_cm = pair_stats(merged_friend, "human", "friend1")

    # Friend vs Gemini (50 overlap)
    merged_fg = friend.merge(gemini, on="id", how="inner")
    fg_n, fg_agree, fg_kappa, fg_cm = pair_stats(merged_fg, "friend1", "gemini")

    # Friend vs ChatGPT (friend only has 50, chatgpt only has 93—find overlap)
    merged_fc = friend.merge(chatgpt, on="id", how="inner")
    fc_n, fc_agree, fc_kappa, fc_cm = pair_stats(merged_fc, "friend1", "chatgpt")

    # Human vs Friend 2 (50 overlap)
    merged_friend2 = human.merge(friend2, on="id", how="inner")
    hf2_n, hf2_agree, hf2_kappa, hf2_cm = pair_stats(merged_friend2, "human", "friend2")

    # Friend 1 vs Friend 2 (50 overlap)
    merged_f1f2 = friend.merge(friend2, on="id", how="inner")
    f1f2_n, f1f2_agree, f1f2_kappa, f1f2_cm = pair_stats(merged_f1f2, "friend1", "friend2")
    
    # Friend 2 vs Gemini (50 overlap)
    merged_f2g = friend2.merge(gemini, on="id", how="inner")
    f2g_n, f2g_agree, f2g_kappa, f2g_cm = pair_stats(merged_f2g, "friend2", "gemini")

    # Friend 3 stats
    merged_friend3 = human.merge(friend3, on="id", how="inner")
    hf3_n, hf3_agree, hf3_kappa, hf3_cm = pair_stats(merged_friend3, "human", "friend3")

    merged_f1f3 = friend.merge(friend3, on="id", how="inner")
    f1f3_n, f1f3_agree, f1f3_kappa, f1f3_cm = pair_stats(merged_f1f3, "friend1", "friend3")

    merged_f2f3 = friend2.merge(friend3, on="id", how="inner")
    f2f3_n, f2f3_agree, f2f3_kappa, f2f3_cm = pair_stats(merged_f2f3, "friend2", "friend3")

    merged_f3g = friend3.merge(gemini, on="id", how="inner")
    f3g_n, f3g_agree, f3g_kappa, f3g_cm = pair_stats(merged_f3g, "friend3", "gemini")

    # 3-way agreement
    three_way["all_agree"] = (three_way["human"] == three_way["chatgpt"]) & (three_way["human"] == three_way["gemini"])
    three_agree = three_way["all_agree"].sum()

    def kappa_interp(k):
        if k < 0: return "Less than chance"
        elif k < 0.21: return "Slight"
        elif k < 0.41: return "Fair"
        elif k < 0.61: return "Moderate"
        elif k < 0.81: return "Substantial"
        else: return "Almost perfect"

    # ══════════════════════════════════════════════
    # BUILD WORKBOOK
    # ══════════════════════════════════════════════
    wb = Workbook()

    # ── SHEET 1: Summary Dashboard ──
    ws1 = wb.active; ws1.title = "Summary"; ws1.sheet_properties.tabColor = "1F4E78"
    write_title(ws1, "Inter-Rater Reliability: Human vs. ChatGPT vs. Gemini")
    ws1.merge_cells('A2:H2')
    ws1['A2'].value = "Pragmatic Function Classification of 'Fuck' in Reddit Comments"
    ws1['A2'].font = Font(color='666666', size=11, name='Arial', italic=True)

    # ── Pairwise Kappa Cards ──
    r = 4
    write_subtitle(ws1, "Pairwise Cohen's Kappa (κ)", r); r += 1
    card_headers = ["Pair", "Samples", "Agreement", "κ", "Interpretation"]
    for ci, h in enumerate(card_headers, 1): ws1.cell(row=r, column=ci).value = h
    hdr(ws1, r, len(card_headers)); r += 1

    pairs = [
        ("Principal Annotator vs. Gemini 3.1 Pro", hg_n, hg_agree, hg_kappa),
        ("Principal Annotator vs. Friend 1", hf_n, hf_agree, hf_kappa),
        ("Principal Annotator vs. Friend 2", hf2_n, hf2_agree, hf2_kappa),
        ("Principal Annotator vs. Friend 3", hf3_n, hf3_agree, hf3_kappa),
        ("Friend 1 vs. Friend 2", f1f2_n, f1f2_agree, f1f2_kappa),
        ("Friend 1 vs. Friend 3", f1f3_n, f1f3_agree, f1f3_kappa),
        ("Friend 2 vs. Friend 3", f2f3_n, f2f3_agree, f2f3_kappa),
        ("Principal Annotator vs. ChatGPT 5.4", hc_n, hc_agree, hc_kappa),
        ("Gemini 3.1 Pro vs. ChatGPT 5.4", gc_n, gc_agree, gc_kappa),
        ("Friend 1 vs. Gemini 3.1 Pro", fg_n, fg_agree, fg_kappa),
        ("Friend 2 vs. Gemini 3.1 Pro", f2g_n, f2g_agree, f2g_kappa),
        ("Friend 3 vs. Gemini 3.1 Pro", f3g_n, f3g_agree, f3g_kappa),
    ]
    for i, (name, n, ag, kp) in enumerate(pairs):
        ws1.cell(row=r, column=1).value = name
        ws1.cell(row=r, column=2).value = n
        ws1.cell(row=r, column=3).value = f"{ag}/{n} ({ag/n*100:.1f}%)"
        ws1.cell(row=r, column=4).value = f"{kp:.3f}"
        ws1.cell(row=r, column=5).value = kappa_interp(kp)
        for ci in range(1, 6):
            c = ws1.cell(row=r, column=ci); dcell(c, i)
            if ci == 1: c.font = BOLD
            if ci == 4: c.font = Font(size=14, name='Arial', bold=True, color='1F4E78')
        r += 1

    # 3-way stat
    r += 1
    write_subtitle(ws1, "Three-Way Agreement (all 3 annotators agree)", r); r += 1
    ws1.cell(row=r, column=1).value = "3-way overlap samples"
    ws1.cell(row=r, column=2).value = len(three_way)
    ws1.cell(row=r, column=1).font = BOLD; ws1.cell(row=r, column=2).font = BIG_NUM
    ws1.cell(row=r, column=1).border = BORDER; ws1.cell(row=r, column=2).border = BORDER
    r += 1
    ws1.cell(row=r, column=1).value = "All 3 agree"
    ws1.cell(row=r, column=2).value = f"{three_agree}/{len(three_way)} ({three_agree/len(three_way)*100:.1f}%)"
    ws1.cell(row=r, column=1).font = BOLD; ws1.cell(row=r, column=2).font = BIG_NUM
    ws1.cell(row=r, column=1).border = BORDER; ws1.cell(row=r, column=2).border = BORDER

    # ── Label Distribution Comparison ──
    r += 2
    write_subtitle(ws1, "Label Distribution Comparison", r); r += 1
    dist_headers = ["Label", f"Principal Annotator (n={len(merged_all)})", f"Gemini 3.1 Pro (n={len(merged_all)})", f"ChatGPT 5.4 (n={len(three_way)})", "P.A. %", "Gemini %", "ChatGPT %"]
    for ci, h in enumerate(dist_headers, 1): ws1.cell(row=r, column=ci).value = h
    hdr(ws1, r, len(dist_headers)); r += 1

    for i, lbl in enumerate(labels):
        h_n = (merged_all["human"] == lbl).sum()
        g_n = (merged_all["gemini"] == lbl).sum()
        c_n = (three_way["chatgpt"] == lbl).sum()
        ws1.cell(row=r, column=1).value = lbl
        ws1.cell(row=r, column=2).value = h_n
        ws1.cell(row=r, column=3).value = g_n
        ws1.cell(row=r, column=4).value = c_n
        ws1.cell(row=r, column=5).value = f"{h_n/len(merged_all)*100:.1f}%"
        ws1.cell(row=r, column=6).value = f"{g_n/len(merged_all)*100:.1f}%"
        ws1.cell(row=r, column=7).value = f"{c_n/len(three_way)*100:.1f}%"
        for ci in range(1, 8):
            c = ws1.cell(row=r, column=ci); dcell(c, i)
            if ci == 1: lcol(c, lbl); c.font = BOLD
        r += 1

    # ── Tie strength breakdown ──
    r += 1
    write_subtitle(ws1, "Agreement by Network Tie Strength (P.A. vs Gemini 3.1 Pro)", r); r += 1
    merged_all["hg_agree"] = merged_all["human"] == merged_all["gemini"]
    tie_h = ["Tie Strength", "Agree", "Total", "Agreement %"]
    for ci, h in enumerate(tie_h, 1): ws1.cell(row=r, column=ci).value = h
    hdr(ws1, r, len(tie_h)); r += 1
    for i, (tie, grp) in enumerate(merged_all.groupby("network_tie_strength")):
        ag = grp["hg_agree"].sum(); tot = len(grp)
        ws1.cell(row=r, column=1).value = tie
        ws1.cell(row=r, column=2).value = int(ag)
        ws1.cell(row=r, column=3).value = tot
        ws1.cell(row=r, column=4).value = f"{ag/tot*100:.1f}%"
        for ci in range(1, 5): dcell(ws1.cell(row=r, column=ci), i)
        r += 1

    for ci in range(1, 9): ws1.column_dimensions[get_column_letter(ci)].width = 18

    # ── SHEET 2: Confusion Matrices ──
    ws2 = wb.create_sheet("Confusion Matrices"); ws2.sheet_properties.tabColor = "2E75B6"
    write_title(ws2, "Pairwise Confusion Matrices")

    r = 3
    r = write_cm(ws2, r, hg_cm, labels, f"Principal Annotator vs. Gemini 3.1 Pro (n={hg_n})", "P.A.", "Gemini")
    r = write_cm(ws2, r + 1, hf_cm, labels, f"Principal Annotator vs. Friend 1 (n={hf_n})", "P.A.", "Friend 1")
    r = write_cm(ws2, r + 1, hf2_cm, labels, f"Principal Annotator vs. Friend 2 (n={hf2_n})", "P.A.", "Friend 2")
    r = write_cm(ws2, r + 1, hf3_cm, labels, f"Principal Annotator vs. Friend 3 (n={hf3_n})", "P.A.", "Friend 3")
    r = write_cm(ws2, r + 1, f1f2_cm, labels, f"Friend 1 vs. Friend 2 (n={f1f2_n})", "Friend 1", "Friend 2")
    r = write_cm(ws2, r + 1, f1f3_cm, labels, f"Friend 1 vs. Friend 3 (n={f1f3_n})", "Friend 1", "Friend 3")
    r = write_cm(ws2, r + 1, f2f3_cm, labels, f"Friend 2 vs. Friend 3 (n={f2f3_n})", "Friend 2", "Friend 3")
    r = write_cm(ws2, r + 1, hc_cm, labels, f"Principal Annotator vs. ChatGPT 5.4 (n={hc_n})", "P.A.", "ChatGPT")
    r = write_cm(ws2, r + 1, gc_cm, labels, f"Gemini vs. ChatGPT (n={gc_n})", "Gemini", "ChatGPT")
    r = write_cm(ws2, r + 1, fg_cm, labels, f"Friend 1 vs. Gemini (n={fg_n})", "Friend 1", "Gemini")
    r = write_cm(ws2, r + 1, f2g_cm, labels, f"Friend 2 vs. Gemini (n={f2g_n})", "Friend 2", "Gemini")
    r = write_cm(ws2, r + 1, f3g_cm, labels, f"Friend 3 vs. Gemini (n={f3g_n})", "Friend 3", "Gemini")

    for ci in range(1, 8): ws2.column_dimensions[get_column_letter(ci)].width = 16

    # ── SHEET 3: Full 3-Way Comparison ──
    ws3 = wb.create_sheet("3-Way Comparison"); ws3.sheet_properties.tabColor = "4CAF50"
    write_title(ws3, f"Three-Way Row-by-Row Comparison (n={len(three_way)})", merge_to='J')

    r = 3
    cols3 = ["ID", "Subreddit", "Tie", "Parent Text", "Comment", "Human", "Gemini", "ChatGPT", "All Agree", "Majority Label"]
    for ci, h in enumerate(cols3, 1): ws3.cell(row=r, column=ci).value = h
    hdr(ws3, r, len(cols3)); r += 1

    widths3 = {1:12, 2:16, 3:12, 4:45, 5:45, 6:13, 7:13, 8:13, 9:10, 10:14}
    for ci, w in widths3.items(): ws3.column_dimensions[get_column_letter(ci)].width = w

    for i, (_, row_data) in enumerate(three_way.iterrows()):
        # Compute majority label
        vote_labels = [row_data["human"], row_data["gemini"], row_data["chatgpt"]]
        from collections import Counter
        counts = Counter(vote_labels)
        majority = counts.most_common(1)[0][0] if counts.most_common(1)[0][1] >= 2 else "no majority"

        ws3.cell(row=r, column=1).value = row_data["id"]
        ws3.cell(row=r, column=2).value = row_data["subreddit"]
        ws3.cell(row=r, column=3).value = row_data["network_tie_strength"]
        ws3.cell(row=r, column=4).value = str(row_data["parent_text"])[:250]
        ws3.cell(row=r, column=5).value = str(row_data["text"])[:250]
        ws3.cell(row=r, column=6).value = row_data["human"]
        ws3.cell(row=r, column=7).value = row_data["gemini"]
        ws3.cell(row=r, column=8).value = row_data["chatgpt"]
        ws3.cell(row=r, column=9).value = "✓" if row_data["all_agree"] else "✗"
        ws3.cell(row=r, column=10).value = majority

        for ci in range(1, 11):
            c = ws3.cell(row=r, column=ci)
            align = LEFT_WRAP if ci in [4, 5] else CENTER
            dcell(c, i, align)

        # Color label cells
        for ci in [6, 7, 8, 10]: lcol(ws3.cell(row=r, column=ci), ws3.cell(row=r, column=ci).value)

        # Agreement indicator
        ac = ws3.cell(row=r, column=9)
        if row_data["all_agree"]:
            ac.fill = LIGHT_GREEN; ac.font = Font(size=12, name='Arial', bold=True, color='2E7D32')
        else:
            ac.fill = LIGHT_RED; ac.font = Font(size=12, name='Arial', bold=True, color='C62828')
        r += 1

    ws3.freeze_panes = 'A4'

    # ── SHEET 4: Human vs Gemini Full ──
    ws4 = wb.create_sheet("Human vs Gemini (Full)"); ws4.sheet_properties.tabColor = "FF9800"
    write_title(ws4, f"Human vs. Gemini — Full Comparison (n={len(merged_all)})")

    r = 3
    cols4 = ["ID", "Subreddit", "Tie", "Parent Text", "Comment", "Human", "Gemini", "Agreement"]
    for ci, h in enumerate(cols4, 1): ws4.cell(row=r, column=ci).value = h
    hdr(ws4, r, len(cols4)); r += 1

    widths4 = {1:12, 2:16, 3:12, 4:45, 5:45, 6:13, 7:13, 8:10}
    for ci, w in widths4.items(): ws4.column_dimensions[get_column_letter(ci)].width = w

    for i, (_, row_data) in enumerate(merged_all.iterrows()):
        ws4.cell(row=r, column=1).value = row_data["id"]
        ws4.cell(row=r, column=2).value = row_data["subreddit"]
        ws4.cell(row=r, column=3).value = row_data["network_tie_strength"]
        ws4.cell(row=r, column=4).value = str(row_data["parent_text"])[:250]
        ws4.cell(row=r, column=5).value = str(row_data["text"])[:250]
        ws4.cell(row=r, column=6).value = row_data["human"]
        ws4.cell(row=r, column=7).value = row_data["gemini"]
        ws4.cell(row=r, column=8).value = "✓" if row_data["hg_agree"] else "✗"

        for ci in range(1, 9):
            c = ws4.cell(row=r, column=ci)
            align = LEFT_WRAP if ci in [4, 5] else CENTER
            dcell(c, i, align)

        lcol(ws4.cell(row=r, column=6), row_data["human"])
        lcol(ws4.cell(row=r, column=7), row_data["gemini"])

        ac = ws4.cell(row=r, column=8)
        if row_data["hg_agree"]:
            ac.fill = LIGHT_GREEN; ac.font = Font(size=12, name='Arial', bold=True, color='2E7D32')
        else:
            ac.fill = LIGHT_RED; ac.font = Font(size=12, name='Arial', bold=True, color='C62828')
        r += 1

    ws4.freeze_panes = 'A4'

    # ── SHEET 5: Gemini vs ChatGPT ──
    three_way["gc_agree"] = three_way["gemini"] == three_way["chatgpt"]
    ws_gc = wb.create_sheet("Gemini vs ChatGPT"); ws_gc.sheet_properties.tabColor = "9C27B0"
    write_title(ws_gc, f"Gemini vs. ChatGPT — Comparison (n={len(three_way)})")

    r = 3
    cols_gc = ["ID", "Subreddit", "Tie", "Parent Text", "Comment", "Gemini", "ChatGPT", "Agreement"]
    for ci, h in enumerate(cols_gc, 1): ws_gc.cell(row=r, column=ci).value = h
    hdr(ws_gc, r, len(cols_gc)); r += 1

    for ci, w in widths4.items(): ws_gc.column_dimensions[get_column_letter(ci)].width = w

    for i, (_, row_data) in enumerate(three_way.iterrows()):
        ws_gc.cell(row=r, column=1).value = row_data["id"]
        ws_gc.cell(row=r, column=2).value = row_data["subreddit"]
        ws_gc.cell(row=r, column=3).value = row_data["network_tie_strength"]
        ws_gc.cell(row=r, column=4).value = str(row_data["parent_text"])[:250]
        ws_gc.cell(row=r, column=5).value = str(row_data["text"])[:250]
        ws_gc.cell(row=r, column=6).value = row_data["gemini"]
        ws_gc.cell(row=r, column=7).value = row_data["chatgpt"]
        ws_gc.cell(row=r, column=8).value = "✓" if row_data["gc_agree"] else "✗"

        for ci in range(1, 9):
            c = ws_gc.cell(row=r, column=ci)
            align = LEFT_WRAP if ci in [4, 5] else CENTER
            dcell(c, i, align)

        lcol(ws_gc.cell(row=r, column=6), row_data["gemini"])
        lcol(ws_gc.cell(row=r, column=7), row_data["chatgpt"])

        ac = ws_gc.cell(row=r, column=8)
        if row_data["gc_agree"]:
            ac.fill = LIGHT_GREEN; ac.font = Font(size=12, name='Arial', bold=True, color='2E7D32')
        else:
            ac.fill = LIGHT_RED; ac.font = Font(size=12, name='Arial', bold=True, color='C62828')
        r += 1

    ws_gc.freeze_panes = 'A4'

    # ── SHEET 6: Disagreements ──
    ws5 = wb.create_sheet("Disagreements"); ws5.sheet_properties.tabColor = "F44336"

    # All rows where at least one AI disagrees with human
    disagree = merged_all[merged_all["human"] != merged_all["gemini"]].copy()
    write_title(ws5, f"Human–Gemini Disagreements ({len(disagree)} cases)")

    r = 3
    cols5 = ["ID", "Subreddit", "Tie", "Parent Text", "Comment", "Human", "Gemini", "Shift"]
    for ci, h in enumerate(cols5, 1): ws5.cell(row=r, column=ci).value = h
    hdr(ws5, r, len(cols5), DARK_RED); r += 1

    for ci, w in widths4.items(): ws5.column_dimensions[get_column_letter(ci)].width = w

    for i, (_, row_data) in enumerate(disagree.iterrows()):
        ws5.cell(row=r, column=1).value = row_data["id"]
        ws5.cell(row=r, column=2).value = row_data["subreddit"]
        ws5.cell(row=r, column=3).value = row_data["network_tie_strength"]
        ws5.cell(row=r, column=4).value = str(row_data["parent_text"])[:250]
        ws5.cell(row=r, column=5).value = str(row_data["text"])[:250]
        ws5.cell(row=r, column=6).value = row_data["human"]
        ws5.cell(row=r, column=7).value = row_data["gemini"]
        ws5.cell(row=r, column=8).value = f"{row_data['human']} → {row_data['gemini']}"

        for ci in range(1, 9):
            c = ws5.cell(row=r, column=ci)
            align = LEFT_WRAP if ci in [4, 5] else CENTER
            dcell(c, i, align)

        lcol(ws5.cell(row=r, column=6), row_data["human"])
        lcol(ws5.cell(row=r, column=7), row_data["gemini"])
        r += 1

    ws5.freeze_panes = 'A4'

    # ── Save ──
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)

    # ── Print Summary ──
    print("=" * 65)
    print("  4-WAY INTER-RATER RELIABILITY REPORT")
    print("=" * 65)
    print()
    print(f"  Human vs. Gemini    κ = {hg_kappa:.3f}  ({kappa_interp(hg_kappa)})  n={hg_n}  agree={hg_agree} ({hg_agree/hg_n*100:.1f}%)")
    print(f"  Human vs. Friend 1  κ = {hf_kappa:.3f}  ({kappa_interp(hf_kappa)})  n={hf_n}  agree={hf_agree} ({hf_agree/hf_n*100:.1f}%)")
    print(f"  Human vs. Friend 2  κ = {hf2_kappa:.3f}  ({kappa_interp(hf2_kappa)})  n={hf2_n}  agree={hf2_agree} ({hf2_agree/hf2_n*100:.1f}%)")
    print(f"  Human vs. Friend 3  κ = {hf3_kappa:.3f}  ({kappa_interp(hf3_kappa)})  n={hf3_n}  agree={hf3_agree} ({hf3_agree/hf3_n*100:.1f}%)")
    print(f"  Friend 1 vs Friend 2 κ = {f1f2_kappa:.3f}  ({kappa_interp(f1f2_kappa)})  n={f1f2_n}  agree={f1f2_agree} ({f1f2_agree/f1f2_n*100:.1f}%)")
    print(f"  Friend 1 vs Friend 3 κ = {f1f3_kappa:.3f}  ({kappa_interp(f1f3_kappa)})  n={f1f3_n}  agree={f1f3_agree} ({f1f3_agree/f1f3_n*100:.1f}%)")
    print(f"  Friend 2 vs Friend 3 κ = {f2f3_kappa:.3f}  ({kappa_interp(f2f3_kappa)})  n={f2f3_n}  agree={f2f3_agree} ({f2f3_agree/f2f3_n*100:.1f}%)")
    print(f"  Human vs. ChatGPT   κ = {hc_kappa:.3f}  ({kappa_interp(hc_kappa)})  n={hc_n}  agree={hc_agree} ({hc_agree/hc_n*100:.1f}%)")
    print(f"  Gemini vs. ChatGPT  κ = {gc_kappa:.3f}  ({kappa_interp(gc_kappa)})  n={gc_n}  agree={gc_agree} ({gc_agree/gc_n*100:.1f}%)")
    print(f"  Friend 1 vs Gemini  κ = {fg_kappa:.3f}  ({kappa_interp(fg_kappa)})  n={fg_n}  agree={fg_agree} ({fg_agree/fg_n*100:.1f}%)")
    print(f"  Friend 2 vs Gemini  κ = {f2g_kappa:.3f}  ({kappa_interp(f2g_kappa)})  n={f2g_n}  agree={f2g_agree} ({f2g_agree/f2g_n*100:.1f}%)")
    print(f"  Friend 3 vs Gemini  κ = {f3g_kappa:.3f}  ({kappa_interp(f3g_kappa)})  n={f3g_n}  agree={f3g_agree} ({f3g_agree/f3g_n*100:.1f}%)")
    print()
    print(f"  3-way full agreement: {three_agree}/{len(three_way)} ({three_agree/len(three_way)*100:.1f}%)")
    print()
    print(f"  ✓ Saved: {OUTPUT_FILE}")
    print(f"  Sheets: Summary | Confusion Matrices | 3-Way Comparison | Human vs Gemini (Full) | Disagreements")

if __name__ == "__main__":
    main()
