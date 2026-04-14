import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = "data/labeled/to_label.tsv"
OUTPUT_FILE = "data/labeled/to_label.xlsx"

def main():
    # Load the TSV file
    df = pd.read_csv(INPUT_FILE, sep="\t")
    
    # Create a Pandas Excel writer using openpyxl as the engine
    writer = pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl')
    
    # Convert the dataframe to an XlsxWriter object
    df.to_excel(writer, index=False, sheet_name='Labeling Task')
    
    # Get the workbook and worksheet objects
    workbook = writer.book
    worksheet = workbook['Labeling Task']
    
    # Styling Constants
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Dark Blue
    header_font = Font(color='FFFFFF', bold=True, size=12, name='Arial')
    odd_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid') # Light Grey
    border_side = Side(style='thin', color='BFBFBF')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Define Column Widths
    # id, subreddit, network_tie_strength, parent_text, text, label
    widths = {
        1: 12, # id
        2: 20, # subreddit
        3: 20, # network_tie_strength
        4: 60, # parent_text
        5: 60, # text
        6: 15  # label
    }
    
    # Apply Styling to Header
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        
        # Set Width
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = widths.get(col_num, 15)
        
    # Apply Styling to Rows
    for row_num in range(2, len(df) + 2):
        is_odd = row_num % 2 != 0
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            
            # Text alignment
            if col_num in [4, 5]: # parent_text, text
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='top')
            
            # Zebra stripes
            if is_odd:
                cell.fill = odd_row_fill
            
            # Borders
            cell.border = border
            
            # Font
            cell.font = Font(size=11, name='Arial')
            
    # Add a data validation list for the label column (Step 6)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"aggression,bonding,emphasis,frustration,ambiguous"', allow_blank=True)
    dv.error ='Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select from the list'
    dv.promptTitle = 'Select Label'
    worksheet.add_data_validation(dv)
    
    # Apply validation to the whole label column (excluding header)
    label_col_letter = get_column_letter(6)
    dv.add(f"{label_col_letter}2:{label_col_letter}{len(df)+1}")
    
    # Freeze the top row
    worksheet.freeze_panes = 'A2'
    
    # Save the file
    writer.close()
    print(f"Successfully created formatted academic Excel file: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
